# civitai_scraper.py
# 爬取 civitai.red 集合页的所有模型，提取提示词，归档到 Excel
# 依赖: pip install requests openpyxl playwright
# 运行前: playwright install chromium

import requests
import json
import time
import re
import os
from collections import Counter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from playwright.sync_api import sync_playwright

# ============================================================
# ✅ 配置区（你只需要改这里）
# ============================================================
CONFIG = {
    # 你的 civitai API key（备用，Bookmark 模式不需要）
    "api_key": "cccf8d6074caec2097a908516ad3ba66",

    # 你的 Bookmark 页面地址
    # civitai.red 一般是 https://civitai.red/user/{你的用户名}/bookmarks
    "collection_url": "https://civitai.red/user/louis748800932/bookmarks",

    # 输出文件名
    "output_file": f"civitai_prompts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",

    # 请求间隔（秒），防止被封
    "request_delay": 1.5,

    # 最大爬取数量（0 = 不限制）
    "max_models": 0,
}

# ============================================================
# 颜色主题（Excel 用）
# ============================================================
COLORS = {
    "character": {"header": "FF6B9D", "row_odd": "FFE4EF", "row_even": "FFF0F6"},
    "pose":      {"header": "6B9DFF", "row_odd": "E4EEFF", "row_even": "F0F4FF"},
    "concept":   {"header": "6BFF9D", "row_odd": "E4FFE4", "row_even": "F0FFF4"},
    "other":     {"header": "FFCC6B", "row_odd": "FFF3E4", "row_even": "FFFAF0"},
}

# ============================================================
# 第一步：通过 Playwright 登录并爬取 Bookmarked Models
# ============================================================
def fetch_bookmarked_models(config: dict) -> list:
    """
    通过 Playwright 浏览器：
    1. 打开 civitai.red 让用户手动登录
    2. 跳转到 bookmarks 页面
    3. 拦截 tRPC/API 响应提取模型数据
    """
    models = []
    captured_data = []

    print("🌐 启动浏览器...")
    print("📌 请在浏览器中手动登录你的 Civitai 账号")
    print("   登录完成后，脚本会自动继续爬取\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled"]
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        )
        page = context.new_page()

        # ---- 拦截所有包含 model 数据的 API 响应 ----
        def handle_response(response):
            url = response.url
            if response.status != 200:
                return
            try:
                data = response.json()
            except Exception:
                return

            # 拦截各种可能的 API 格式
            if "items" in data and isinstance(data["items"], list):
                for item in data["items"]:
                    if isinstance(item, dict) and "id" in item and "name" in item:
                        # 避免重复
                        existing_ids = {m.get("id") for m in captured_data}
                        if item["id"] not in existing_ids:
                            captured_data.append(item)
                print(f"📦 拦截到 {len(data['items'])} 个模型 (累计 {len(captured_data)})")

        page.on("response", handle_response)

        # ---- 打开 civitai.red 登录页 ----
        print("🔗 打开 civitai.red ...")
        page.goto("https://civitai.red/login", wait_until="domcontentloaded", timeout=30000)

        # ---- 等待用户登录（检测到 avatar 菜单出现即登录成功） ----
        print("⏳ 等待登录... (请在浏览器中完成登录)")
        try:
            # 等待用户头像菜单出现，最多等 5 分钟
            page.wait_for_selector(
                "button[aria-label*='profile'], [data-testid='user-menu'], .avatar, nav a[href*='/user/']",
                timeout=300000  # 5 分钟
            )
            print("✅ 登录成功！\n")
        except Exception:
            print("⚠️ 未检测到登录状态，尝试继续...")

        # ---- 跳转到 Bookmarks 页面 ----
        bookmark_url = config.get("collection_url", "https://civitai.red/user/bookmarks")
        print(f"🔗 跳转到 Bookmarks: {bookmark_url}")
        page.goto(bookmark_url, wait_until="networkidle", timeout=60000)
        time.sleep(2)

        # ---- 滚动加载所有模型 ----
        print("⏬ 滚动加载所有收藏模型...")
        last_count = 0
        no_change_rounds = 0
        max_rounds = 60

        for _ in range(max_rounds):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(1.5)

            current = len(captured_data)
            if current == last_count:
                no_change_rounds += 1
            else:
                no_change_rounds = 0
                last_count = current
                print(f"📊 已捕获 {current} 个模型")

            if no_change_rounds >= 5:
                print("✅ 无新数据，加载完毕")
                break

            if config["max_models"] > 0 and current >= config["max_models"]:
                print(f"⚡ 达到上限 {config['max_models']}")
                break

        models = captured_data
        print(f"\n✅ 共捕获 {len(models)} 个 Bookmarked 模型")
        browser.close()

    return models


# ============================================================
# 第三步：从单个模型数据中提取提示词
# ============================================================
def extract_prompts_from_model(model: dict, config: dict) -> dict:
    """
    从模型数据中提取：
    - trigger words（触发词）
    - 示例图片的 positive prompt
    - 示例图片的 negative prompt
    - 模型分类 (character / pose / concept / other)
    """
    model_id = model.get("id", "")
    model_name = model.get("name", "未知")
    model_type = model.get("type", "").upper()  # LORA / LOCON 等

    # ---- 分类判断 ----
    # 根据 tags 或 name 判断是 character / pose / concept
    tags = [t.get("name", "").lower() for t in model.get("tags", [])]
    category = classify_model(model_name, tags)

    # ---- 提取 Trigger Words ----
    trigger_words = []
    model_versions = model.get("modelVersions", [])
    if model_versions:
        # 取最新版本
        latest = model_versions[0]
        trigger_words = latest.get("trainedWords", [])

    # ---- 提取示例图片的 Prompt ----
    positive_prompts = []
    negative_prompts = []

    for version in model_versions[:1]:  # 只取最新版本的示例图
        images = version.get("images", [])
        for img in images[:5]:  # 最多取5张图的提示词
            meta = img.get("meta", {})
            if not meta:
                continue
            pos = meta.get("prompt", "") or meta.get("Prompt", "")
            neg = meta.get("negativePrompt", "") or meta.get("Negative prompt", "")
            if pos:
                positive_prompts.append(pos.strip())
            if neg:
                negative_prompts.append(neg.strip())

    # ---- 模型 URL ----
    model_url = f"https://civitai.red/models/{model_id}"

    # ---- 描述 ----
    description = model.get("description", "") or ""
    # 去掉 HTML 标签
    description = re.sub(r"<[^>]+>", "", description).strip()
    description = description[:500] if len(description) > 500 else description

    return {
        "id": model_id,
        "name": model_name,
        "type": model_type,
        "category": category,
        "tags": ", ".join(tags[:10]),
        "trigger_words": ", ".join(trigger_words),
        "positive_prompt_samples": "\n---\n".join(positive_prompts[:3]),
        "negative_prompt_samples": "\n---\n".join(negative_prompts[:3]),
        "description": description,
        "url": model_url,
        "downloads": model.get("stats", {}).get("downloadCount", 0),
        "rating": model.get("stats", {}).get("rating", 0),
    }


# ============================================================
# 第四步：模型分类逻辑
# ============================================================
def classify_model(name: str, tags: list) -> str:
    """
    根据模型名称和标签判断分类
    返回: character / pose / concept / other
    """
    name_lower = name.lower()
    tags_str = " ".join(tags)

    # Character 判断关键词
    char_keywords = [
        "character", "girl", "boy", "woman", "man", "person",
        "oc", "original character", "anime", "waifu", "husbando",
        "chara", "人物", "角色",
    ]
    # Pose 判断关键词
    pose_keywords = [
        "pose", "position", "action", "gesture", "sitting",
        "standing", "lying", "kneeling", "动作", "姿势",
    ]
    # Concept 判断关键词
    concept_keywords = [
        "style", "concept", "art style", "aesthetic", "lighting",
        "background", "scene", "environment", "texture", "effect",
        "风格", "概念",
    ]

    # 优先检查 tags
    for kw in char_keywords:
        if kw in tags_str:
            return "character"
    for kw in pose_keywords:
        if kw in tags_str:
            return "pose"
    for kw in concept_keywords:
        if kw in tags_str:
            return "concept"

    # 再检查名字
    for kw in char_keywords:
        if kw in name_lower:
            return "character"
    for kw in pose_keywords:
        if kw in name_lower:
            return "pose"
    for kw in concept_keywords:
        if kw in name_lower:
            return "concept"

    return "other"


# ============================================================
# 第五步：写入 Excel
# ============================================================
def write_to_excel(all_data: list, output_file: str):
    """
    将爬取结果写入 Excel，按分类分 Sheet
    每个 Sheet 都有漂亮的表头和交替行颜色
    """
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认空 Sheet

    # 按分类分组
    categories = {"character": [], "pose": [], "concept": [], "other": []}
    for item in all_data:
        cat = item.get("category", "other")
        categories[cat].append(item)

    # Sheet 名称映射
    sheet_names = {
        "character": "🧑 Character 人物",
        "pose": "🤸 Pose 动作",
        "concept": "💡 Concept 概念",
        "other": "📦 Other 其他",
    }

    # 表头定义
    headers = [
        ("ID", 10),
        ("模型名称 Name", 35),
        ("类型 Type", 10),
        ("标签 Tags", 30),
        ("触发词 Trigger Words", 40),
        ("正向提示词示例\nPositive Prompts", 60),
        ("负向提示词示例\nNegative Prompts", 50),
        ("描述 Description", 50),
        ("下载量", 12),
        ("评分", 10),
        ("链接 URL", 45),
    ]

    data_keys = [
        "id", "name", "type", "tags", "trigger_words",
        "positive_prompt_samples", "negative_prompt_samples",
        "description", "downloads", "rating", "url",
    ]

    for cat, items in categories.items():
        if not items:
            continue

        color = COLORS[cat]
        ws = wb.create_sheet(title=sheet_names[cat])

        # ---- 写表头 ----
        header_fill = PatternFill(
            "solid", fgColor=color["header"]
        )
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, (header_text, col_width) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[1].height = 35

        # ---- 写数据行 ----
        for row_idx, item in enumerate(items, start=2):
            is_odd = (row_idx % 2 == 1)
            row_fill = PatternFill(
                "solid",
                fgColor=color["row_odd"] if is_odd else color["row_even"]
            )
            row_align_wrap = Alignment(
                vertical="top", wrap_text=True
            )
            row_align_center = Alignment(
                horizontal="center", vertical="top"
            )

            for col_idx, key in enumerate(data_keys, start=1):
                value = item.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.fill = row_fill
                cell.border = thin_border
                cell.font = Font(size=10)

                # URL 列做超链接
                if key == "url" and value:
                    cell.hyperlink = value
                    cell.font = Font(
                        size=10, color="0563C1", underline="single"
                    )

                # 数字列居中
                if key in ("id", "downloads", "rating"):
                    cell.alignment = row_align_center
                else:
                    cell.alignment = row_align_wrap

            ws.row_dimensions[row_idx].height = 80

        # ---- 冻结首行 ----
        ws.freeze_panes = "A2"

        # ---- 添加汇总信息到最后 ----
        summary_row = len(items) + 3
        ws.cell(
            row=summary_row, column=1,
            value=f"共 {len(items)} 个{sheet_names[cat]}模型"
        ).font = Font(bold=True, italic=True, color="888888")

        print(f"✅ Sheet [{sheet_names[cat]}] 写入 {len(items)} 条数据")

    # ---- 添加总览 Sheet ----
    ws_summary = wb.create_sheet(title="📊 总览 Summary", index=0)
    write_summary_sheet(ws_summary, categories, sheet_names)

    wb.save(output_file)
    print(f"\n🎉 Excel 文件已保存: {output_file}")


# ============================================================
# 第六步：总览 Sheet
# ============================================================
def write_summary_sheet(ws, categories: dict, sheet_names: dict):
    """写入总览统计 Sheet"""

    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="333333")
    title_align = Alignment(horizontal="center", vertical="center")

    # 标题
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "🕷️ Civitai 模型提示词归档总览"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_align
    ws.row_dimensions[1].height = 40

    # 生成时间
    ws["A2"] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, color="888888")

    # 表头
    headers = ["分类", "英文名", "模型数量", "备注"]
    fills = ["FF6B9D", "6B9DFF", "6BFF9D", "FFCC6B"]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="555555")
        cell.alignment = Alignment(horizontal="center")

    # 数据
    cat_labels = {
        "character": ("🧑 人物", "Character", "角色、人物外观相关"),
        "pose": ("🤸 动作", "Pose/Action", "姿势、动作、体态相关"),
        "concept": ("💡 概念", "Concept/Style", "风格、概念、场景相关"),
        "other": ("📦 其他", "Other", "未分类模型"),
    }

    total = 0
    for row_idx, (cat, (label, eng, note)) in enumerate(
        cat_labels.items(), start=5
    ):
        count = len(categories.get(cat, []))
        total += count
        color = COLORS[cat]["header"]

        ws.cell(row=row_idx, column=1, value=label).fill = PatternFill(
            "solid", fgColor=color
        )
        ws.cell(row=row_idx, column=2, value=eng)
        ws.cell(row=row_idx, column=3, value=count).alignment = Alignment(
            horizontal="center"
        )
        ws.cell(row=row_idx, column=4, value=note)

    # 总计
    ws.cell(row=9, column=1, value="📊 合计").font = Font(bold=True)
    ws.cell(row=9, column=3, value=total).font = Font(bold=True)
    ws.cell(row=9, column=3).alignment = Alignment(horizontal="center")

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.column_dimensions["D"].width = 30


# ============================================================
# 主程序入口
# ============================================================
def main():
    print("=" * 60)
    print("Civitai Bookmarked Models - 提示词爬取归档工具")
    print("=" * 60)

    config = CONFIG

    # 爬取 Bookmarked Models
    models = fetch_bookmarked_models(config)

    if not models:
        print("没有获取到任何模型数据，请确认已登录且 Bookmark 页面有内容")
        return

    print(f"\n共获取 {len(models)} 个模型，开始提取提示词...")

    # 提取提示词
    all_data = []
    for i, model in enumerate(models):
        data = extract_prompts_from_model(model, config)
        all_data.append(data)

        if (i + 1) % 10 == 0:
            print(f"已处理 {i+1}/{len(models)} 个模型...")

    # 统计分类
    cat_count = Counter(d["category"] for d in all_data)
    print(f"\n分类统计:")
    for cat, count in cat_count.items():
        print(f"   {cat}: {count} 个")

    # 写入 Excel
    output_path = Path(config["output_file"])
    write_to_excel(all_data, str(output_path))

    print(f"\n全部完成！文件保存在: {output_path.absolute()}")


if __name__ == "__main__":
    main()
